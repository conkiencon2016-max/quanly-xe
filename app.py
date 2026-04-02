from flask import Flask, render_template, request, redirect, session
import sqlite3
from datetime import datetime, date, timedelta
import requests
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import os
from dotenv import load_dotenv
import time
from openpyxl import Workbook
from openpyxl.styles import Font
from flask import send_file
import io
from zalo_service import handle_message
from io import BytesIO
from collections import defaultdict
import re
import threading

# =========================
# FORMAT DATETIME (GLOBAL)
# =========================
def format_date(val):
    if not val:
        return ""
    try:
        return datetime.fromisoformat(val).strftime("%d/%m/%Y")
    except:
        return val

def format_datetime(val):
    if not val:
        return ""
    try:
        return datetime.fromisoformat(val).strftime("%d/%m/%Y %H:%M")
    except:
        return val

def format_datetime_input(val):
    """dùng cho input datetime-local"""
    if not val:
        return ""
    try:
        return datetime.fromisoformat(val).strftime("%Y-%m-%dT%H:%M")
    except:
        return val

db_lock = threading.Lock()

def execute_retry(con, query, params=(), retries=5):
    for i in range(retries):
        try:
            return con.execute(query, params)
        except sqlite3.OperationalError as e:
            if "locked" in str(e):
                time.sleep(0.2 * (i + 1))
            else:
                raise
    raise Exception("DB bị lock quá lâu")

def ai_parse_command(text):

    text = text.lower()

    result = {
        "intent": None,
        "plate": None,
        "phone": None,
        "content": None
    }

    # ======================
    # điều xe
    # ======================
    if "điều xe" in text or "dieu xe" in text:

        result["intent"] = "dieuxe"

        plate = re.search(r"\d{2}[a-z]-\d{3}\.\d{2}", text)
        phone = re.search(r"\d{9,11}", text)

        if plate:
            result["plate"] = plate.group().upper()

        if phone:
            result["phone"] = phone.group()

        # nội dung sau chữ "đi"
        if "đi" in text:
            content = text.split("đi",1)[1]
            result["content"] = content.strip()

    # ======================
    # xem xe
    # ======================
    if text.startswith("xe "):

        result["intent"] = "xe"

        plate = re.search(r"\d{2}[a-z]-\d{3}\.\d{2}", text)

        if plate:
            result["plate"] = plate.group().upper()

    # ======================
    # tài xế
    # ======================
    if "tài xế" in text or "taixe" in text:

        if "rảnh" in text:
            result["intent"] = "taixeranh"

    # ======================
    # thống kê
    # ======================
    if "thống kê" in text or "thong ke" in text:

        result["intent"] = "thongke"

    return result
# =========================
# INIT APP
# =========================
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "change-this-secret")

# Session timeout 15 phút
app.permanent_session_lifetime = timedelta(minutes=15)

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=False  # True nếu chạy HTTPS
)
TELEGRAM_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
ZALO_BOT_TOKEN = os.getenv("ZALO_BOT_TOKEN")
# =========================
# TELEGRAM SERVICE
# =========================
def send_telegram(chat_id, message):
    if not TELEGRAM_TOKEN or not chat_id:
        return

    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"

    try:
        res = requests.post(url, json={
            "chat_id": chat_id,
            "text": message
        }, timeout=5)

        if res.status_code != 200:
            print("Telegram lỗi:", res.text)

    except Exception as e:
        print("Telegram exception:", e)
# =========================
# KẾT NỐI DATABASE
# =========================
def db():
    con = sqlite3.connect(
        "fleet.db",
        timeout=30,
        check_same_thread=False
    )
    con.row_factory = sqlite3.Row

    # 🔥 QUAN TRỌNG NHẤT
    con.execute("PRAGMA journal_mode=WAL;")
    con.execute("PRAGMA synchronous=NORMAL;")

    return con
# =========================
# sesion logout
# =========================


@app.before_request
def auto_session_timeout():
    session.permanent = True
    now = int(time.time())

    if "last_activity" in session:
        elapsed = now - session["last_activity"]

        # 15 phút = 900 giây
        if elapsed > 900:
            session.clear()
            return redirect("/login")

    session["last_activity"] = now

# =========================
# LOGIN DECORATOR PHÂN QUYỀN
# =========================


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("role") != "admin":
            return "Không có quyền truy cập", 403
        return f(*args, **kwargs)
    return decorated


def driver_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("role") != "driver":
            return "Không có quyền truy cập", 403
        return f(*args, **kwargs)
    return decorated


# =========================
# TRANG CHỦ
# =========================
@app.route("/")
@login_required
def home():
    return render_template(
        "home.html",
        role=session.get("role"),
        username=session.get("username")
    )
# =========================
# ZALO CHO TAI XE (ĐÃ TỐI ƯU HÓA)
# =========================
def gui_zalo_cho_taixe(chat_id, noi_dung):

    BOT_TOKEN = os.getenv("ZALO_BOT_TOKEN")

    url = f"https://bot-api.zaloplatforms.com/bot{BOT_TOKEN}/sendMessage"

    payload = {
        "chat_id": chat_id,
        "text": noi_dung
    }

    headers = {
        "Content-Type": "application/json"
    }

    try:

        r = requests.post(
            url,
            json=payload,
            headers=headers,
            timeout=10
        )

        print("Zalo status:", r.status_code)
        print("Zalo response:", r.text)

        if r.status_code == 200:
            res = r.json()
            return res.get("ok", False)

        return False

    except Exception as e:
        print("❌ Lỗi gửi Zalo:", e)
        return "OK"
# =========================
# telegram CHO TAI XE (ĐÃ TỐI ƯU HÓA)
# =========================

def send_telegram(chat_id, message):

    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"

    payload = {
        "chat_id": chat_id,
        "text": message,
        "parse_mode": "HTML"
    }

    try:

        r = requests.post(url, json=payload, timeout=10)

        print("Telegram status:", r.status_code)
        print("Telegram response:", r.text)

        return r.status_code == 200

    except Exception as e:

        print("Telegram error:", e)
        return "OK"
# =========================
# ĐIỀU XE (WEB CON)
# =========================
@app.route("/dieu-xe")
@login_required
@admin_required
def dieu_xe():
    con = db()

    # =========================
    # AUTO FILL TỪ YÊU CẦU
    # =========================
    auto_id = request.args.get("auto_fill")
    yeu_cau = None

    if auto_id:
        yeu_cau = con.execute("""
            SELECT * FROM yeu_cau_xe WHERE id=?
        """, (auto_id,)).fetchone()

    # =========================
    # DASHBOARD
    # =========================
    tong_xe = con.execute(
        "SELECT COUNT(*) FROM vehicles"
    ).fetchone()[0]

    xe_hoat_dong = con.execute(
        "SELECT COUNT(*) FROM vehicles WHERE status = 1"
    ).fetchone()[0]

    xe_o_nha = con.execute(
        "SELECT COUNT(*) FROM vehicles WHERE status = 0"
    ).fetchone()[0]

    # =========================
    # FILTER
    # =========================
    filter_status = request.args.get("status")

    sql_filter = ""
    params = []

    if filter_status == "active":
        sql_filter = "WHERE vehicles.status = 1"
    elif filter_status == "home":
        sql_filter = "WHERE vehicles.status = 0"

    # =========================
    # LẤY XE
    # =========================
    vehicles_raw = con.execute(f"""
        SELECT
            vehicles.id,
            vehicles.plate,
            vehicles.status,
            drivers.name AS driver_name,
            vehicles.start_time,
            vehicles.end_time,
            vehicles.work_content,
            vehicles.requester
        FROM vehicles
        LEFT JOIN drivers ON vehicles.driver_id = drivers.id
        {sql_filter}
    """, params).fetchall()

    # =========================
    # DRIVER RẢNH
    # =========================
    busy_drivers = con.execute("""
        SELECT driver_id
        FROM vehicles
        WHERE status = 1 AND driver_id IS NOT NULL
    """).fetchall()

    busy_ids = [d["driver_id"] for d in busy_drivers]

    if busy_ids:
        placeholders = ",".join("?" * len(busy_ids))
        drivers = con.execute(
            f"SELECT id, name FROM drivers WHERE id NOT IN ({placeholders})",
            busy_ids
        ).fetchall()
    else:
        drivers = con.execute(
            "SELECT id, name FROM drivers"
        ).fetchall()

    # =========================
    # FORMAT DATA
    # =========================
    vehicles = []

    for v in vehicles_raw:
        start = end = duration = None

        if v["start_time"]:
            start_dt = datetime.fromisoformat(v["start_time"])
            start = format_date(v["start_time"])

        if v["end_time"]:
            end_dt = datetime.fromisoformat(v["end_time"])
            end = format_date(v["end_time"])

        vehicles.append((
            v["id"],
            v["plate"],
            v["status"],
            v["driver_name"],
            start,
            end,
            duration,
            v["work_content"],
            v["requester"]
        ))

    con.close()

    return render_template(
        "dieu_xe.html",
        vehicles=vehicles,
        drivers=drivers,
        tong_xe=tong_xe,
        xe_hoat_dong=xe_hoat_dong,
        xe_o_nha=xe_o_nha,
        yeu_cau=yeu_cau   # 🔥 QUAN TRỌNG
    )
# =========================
# RA BÃI
# =========================
@app.route("/start/<int:vid>", methods=["POST"])
def start(vid):

    con = db()

    try:
        driver_id = int(request.form["driver_id"])
        start_time = request.form["start_time"]
        work_content = request.form["work_content"]
        requester = request.form.get("requester")
        end_time = request.form.get("end_time")
        # 🔥 check xe đang chạy
        # 🔥 check xe đang chạy
        xe = con.execute("""
             SELECT status FROM vehicles WHERE id=?
        """, (vid,)).fetchone()

        if xe["status"] == 1:
            return "Xe đang hoạt động!", 400

            # 🔥 check driver chuẩn
        busy = con.execute("""
            SELECT id, plate FROM vehicles
            WHERE status = 1 
            AND driver_id = ?
            AND start_time IS NOT NULL
            LIMIT 1
        """, (driver_id,)).fetchone()

        if busy:
            return f"Tài xế đang chạy xe {busy['plate']}", 400

            # 🔥 update atomic
        execute_retry(con, """
            UPDATE vehicles
            SET status = 1,
                driver_id = ?,
                start_time = ?,
                end_time = ?,
                work_content = ?,
                requester = ?
            WHERE id = ? AND status = 0
        """, (driver_id, start_time, end_time, work_content, requester, vid))

        if con.total_changes == 0:
            return "Xe vừa bị người khác điều!", 400

        # lấy thông tin xe + tài xế
        info = con.execute("""
            SELECT v.plate,
                   d.name,
                   d.zalo_user_id,
                   d.telegram_chat_id
            FROM vehicles v
            JOIN drivers d ON v.driver_id = d.id
            WHERE v.id = ?
        """, (vid,)).fetchone()

        con.commit()

    except Exception as e:
        con.rollback()
        return f"Lỗi hệ thống: {str(e)}", 500

    finally:
        con.close()

    # =============================
    # GỬI THÔNG BÁO BOT
    # =============================
    start_dt = datetime.fromisoformat(start_time)
    thoi_gian_dep = start_dt.strftime("%H:%M ngày %d/%m/%Y")
    
    if end_time:
        end_dt = datetime.fromisoformat(end_time)
        thoi_gian_dep1 = end_dt.strftime("%H:%M ngày %d/%m/%Y")
    else:
        thoi_gian_dep1 = "Chưa xác định"
    if info:

        noi_dung = f"""
🚗 THÔNG BÁO ĐIỀU XE

Xe: {info['plate']}
Tài xế: {info['name']}
Thời gian đi: {thoi_gian_dep}
Thời gian về: {thoi_gian_dep1}
Nội dung:
{work_content}

Vui lòng thực hiện theo phân công.
"""

        # gửi zalo
        if info["zalo_user_id"]:
            gui_zalo_cho_taixe(
                info["zalo_user_id"],
                noi_dung
            )

        # gửi telegram
        if info["telegram_chat_id"]:
            send_telegram(
                info["telegram_chat_id"],
                noi_dung
            )

        # ghi log
        con_log = db()
        con_log.execute("""
            INSERT INTO zalo_logs
            (driver_id, plate, content, status)
            VALUES (?, ?, ?, ?)
        """, (
            driver_id,
            info["plate"],
            noi_dung,
            "success"
        ))
        con_log.commit()
        con_log.close()

    return redirect("/dieu-xe")
# =========================
# VÀO BÃI (KẾT THÚC ĐIỀU XE)
# =========================
@app.route("/stop/<int:vid>", methods=["POST"])
def stop(vid):
    con = db()

    try:
        end_time = request.form.get("end_time")
        km_travel = request.form.get("km_travel", 0)

        if not end_time:
            return "Thiếu thời gian kết thúc", 400

        # Ép kiểu km
        try:
            km_travel = int(km_travel)
            if km_travel < 0:
                km_travel = 0
        except:
            km_travel = 0

        # 1️⃣ Lấy thông tin chuyến đi
        trip = con.execute("""
            SELECT
                vehicles.id,
                vehicles.plate,
                vehicles.km,
                drivers.name AS driver_name,
                vehicles.start_time,
                vehicles.work_content,
                vehicles.requester
            FROM vehicles
            LEFT JOIN drivers ON vehicles.driver_id = drivers.id
            WHERE vehicles.id = ?
        """, (vid,)).fetchone()

        if not trip:
            return "Không tìm thấy chuyến xe", 404

        if not trip["start_time"]:
            return "Xe chưa có thời gian bắt đầu", 400

        # 2️⃣ Tính thời gian chạy
        start_dt = datetime.fromisoformat(trip["start_time"])
        end_dt = datetime.fromisoformat(end_time)

        if end_dt < start_dt:
            return "Thời gian kết thúc không hợp lệ", 400

        duration_minutes = int((end_dt - start_dt).total_seconds() / 60)

        # 3️⃣ GHI LỊCH SỬ
        execute_retry(con, """

            INSERT INTO trip_history (
                vehicle_id,
                plate,
                driver_name,
                start_time,
                end_time,
                duration_minutes,
                work_content,
                km_travel,
                requester
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            trip["id"],
            trip["plate"],
            trip["driver_name"],
            trip["start_time"],
            end_time,
            duration_minutes,
            trip["work_content"],
            km_travel,
            trip["requester"]
        ))

        # 4️⃣ CỘNG DỒN KM VÀO XE
        current_km = trip["km"] or 0
        new_km = current_km + km_travel

        con.execute("""
            UPDATE vehicles
            SET km = ?
            WHERE id = ?
        """, (new_km, vid))

        # 5️⃣ RESET TRẠNG THÁI XE
        execute_retry(con,"""
            UPDATE vehicles
            SET status = 0,
                driver_id = NULL,
                start_time = NULL,
                end_time = NULL,
                work_content = NULL,
                requester = NULL
            WHERE id = ?
        """, (vid,))

        con.commit()

    except Exception as e:
        con.rollback()
        return f"Lỗi hệ thống: {str(e)}", 500

    finally:
        con.close()

    return redirect("/dieu-xe")



# =========================
# DRIVER KẾT THÚC XE
# =========================
@app.route("/stop-driver/<int:vid>", methods=["POST"])
@login_required
@driver_required
def stop_driver(vid):

    username = session.get("username")
    con = db()

    try:
        # Lấy driver theo phone
        driver = con.execute("""
            SELECT id, name
            FROM drivers
            WHERE phone = ?
        """, (username,)).fetchone()

        if not driver:
            return "Không có quyền thao tác", 403

        trip = con.execute("""
            SELECT *
            FROM vehicles
            WHERE id=? AND driver_id=? AND status=1
        """, (vid, driver["id"])).fetchone()

        if not trip:
            return "Không có quyền thao tác", 403

        end_time = request.form.get("end_time")
        km_travel = int(request.form.get("km_travel", 0))

        start_dt = datetime.fromisoformat(trip["start_time"])
        end_dt = datetime.fromisoformat(end_time)

        if end_dt < start_dt:
            return "Thời gian không hợp lệ", 400

        duration_minutes = int((end_dt - start_dt).total_seconds() / 60)

        # Lưu lịch sử
        con.execute("""
            INSERT INTO trip_history (
                vehicle_id, plate, driver_name,
                start_time, end_time,
                duration_minutes, work_content, km_travel
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            trip["id"],
            trip["plate"],
            driver["name"],
            trip["start_time"],
            end_time,
            duration_minutes,
            trip["work_content"],
            km_travel
        ))

        # Reset xe
        new_km = (trip["km"] or 0) + km_travel

        con.execute("""
            UPDATE vehicles
            SET km=?,
                status=0,
                driver_id=NULL,
                start_time=NULL,
                end_time=NULL,
                work_content=NULL
            WHERE id=?
        """, (new_km, vid))

        con.commit()

    except Exception as e:
        con.rollback()
        return f"Lỗi hệ thống: {str(e)}", 500

    finally:
        con.close()

    # 🔁 Sau khi reset → quay lại trang
    # Vì không còn status=1 nên trang sẽ trống
    return redirect("/dieu-xe-driver")
# =========================
# QUẢN LÝ TÀI XẾ 
# =========================
@app.route("/quan-ly-tai-xe", methods=["GET", "POST"])
@login_required
@admin_required
def quan_ly_tai_xe():
    con = db()

    # =========================
    # THÊM TÀI XẾ
    # =========================
    if request.method == "POST":
        try:
            name = request.form["name"].strip()
            phone = request.form.get("phone", "").strip()
            address = request.form.get("address", "").strip()
            zalo_user_id = request.form.get("zalo_user_id", "").strip()
            telegram_chat_id = request.form.get("telegram_chat_id", "").strip()

            if not name:
                return "Thiếu tên tài xế", 400

            con.execute("""
                INSERT INTO drivers (name, phone, address, zalo_user_id, telegram_chat_id)
                VALUES (?, ?, ?, ?,?)
            """, (name, phone, address, zalo_user_id, telegram_chat_id))

            con.commit()

        except Exception as e:
            con.rollback()
            return f"Lỗi hệ thống: {str(e)}", 500

        finally:
            con.close()

        return redirect("/quan-ly-tai-xe")

    # =========================
    # LẤY DANH SÁCH TÀI XẾ
    # =========================
    drivers = con.execute("""
        SELECT
            d.id,
            d.name,
            d.phone,
            d.address,
            d.zalo_user_id,
            d.telegram_chat_id,
            CASE
                WHEN EXISTS (
                    SELECT 1 FROM vehicles v
                    WHERE v.driver_id = d.id
                    AND v.status = 1
                )
                THEN 1
                ELSE 0
            END AS is_busy
        FROM drivers d
        ORDER BY d.name
    """).fetchall()

    con.close()

    return render_template(
        "quan_ly_tai_xe.html",
        drivers=drivers
    )
# =========================
# Sửa tài xế
# =========================

@app.route("/sua-tai-xe/<int:did>", methods=["GET", "POST"])
def sua_tai_xe(did):
    con = db()

    if request.method == "POST":
        name = request.form["name"]
        phone = request.form["phone"]
        address = request.form["address"]
        zalo_user_id= request.form["zalo_user_id"]
        telegram_chat_id= request.form["telegram_chat_id"]
        con.execute("""
            UPDATE drivers
            SET name=?, phone=?, address=?, zalo_user_id=?, telegram_chat_id=?
            WHERE id=?
        """, (name, phone, address, zalo_user_id, telegram_chat_id, did))
        con.commit()
        return redirect("/quan-ly-tai-xe")

    driver = con.execute(
        "SELECT * FROM drivers WHERE id = ?",
        (did,)
    ).fetchone()

    return render_template("sua_tai_xe.html", driver=driver)

# =========================
# Xóa tài xế
# =========================

@app.route("/xoa-tai-xe/<int:did>")
def xoa_tai_xe(did):
    con = db()

    busy = execute_retry(con, """
        SELECT 1 FROM vehicles
        WHERE status = 1 AND driver_id = ?
    """, (did,)).fetchone()

    if busy:
        return "Không thể xóa: Tài xế đang điều xe", 400

    execute_retry(con, "DELETE FROM drivers WHERE id = ?", (did,))
    con.commit()
    return redirect("/quan-ly-tai-xe")
# =========================
# quản lý phương tiện
# =========================


@app.route("/quan-ly-xe-menu")
def quan_ly_xe_menu():
    return render_template("quan_ly_xe_menu.html")
# =========================
# lịch sử đăng kiểm
# =========================

@app.route("/lich-su-dang-kiem")
@login_required
def lich_su_dang_kiem():

    con = db()

    search = request.args.get("search", "")
    tu_ngay = request.args.get("tu_ngay")
    den_ngay = request.args.get("den_ngay")

    sql = """
        SELECT dk.*, v.plate
        FROM dang_kiem dk
        JOIN vehicles v ON dk.vehicle_id = v.id
        WHERE 1=1
    """

    params = []

    if search:
        sql += " AND (v.plate LIKE ? OR dk.so_dang_ky LIKE ?)"
        params += [f"%{search}%", f"%{search}%"]

    if tu_ngay:
        sql += " AND dk.ngay_dang_ky >= ?"
        params.append(tu_ngay)

    if den_ngay:
        sql += " AND dk.ngay_dang_ky <= ?"
        params.append(den_ngay)

    sql += " ORDER BY dk.ngay_dang_ky DESC"

    data = con.execute(sql, params).fetchall()

    # =========================
    # EXPORT EXCEL
    # =========================
    if request.args.get("export") == "excel":

        wb = Workbook()
        ws = wb.active
        ws.title = "DangKiem"

        headers = [
            "Biển số",
            "Số đăng ký",
            "Loại",
            "Ngày đăng ký",
            "Ngày hết hạn",
            "Trung tâm",
            "Chi phí",
            "Người thực hiện"
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for r in data:
            ws.append([
                r["plate"],
                r["so_dang_ky"],
                r["loai"],
                format_date(r["ngay_dang_ky"]),
                format_date(r["ngay_het_han"]),
                r["trung_tam"],
                r["chi_phi"],
                r["nguoi_thuc_hien"]
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return send_file(
            file_stream,
            download_name="lich_su_dang_kiem.xlsx",
            as_attachment=True
        )

    con.close()

    return render_template(
        "lich_su_dang_kiem.html",
        danh_sach=data
    )

# =========================
# THÊM ĐĂNG KIỂM (CHUẨN NGHIỆP VỤ)
# =========================
@app.route("/them-dang-kiem", methods=["GET", "POST"])
def them_dang_kiem():
    con = db()   # dùng db() để có row_factory

    if request.method == "POST":
        try:
            vehicle_id = request.form["vehicle_id"]
            so_dang_ky = request.form["so_dang_ky"].strip()
            loai = request.form["loai"]
            ngay_dang_ky = request.form["ngay_dang_ky"]
            ngay_het_han = request.form["ngay_het_han"]
            trung_tam = request.form.get("trung_tam", "").strip()
            chi_phi = request.form.get("chi_phi") or 0
            nguoi_thuc_hien = request.form.get("nguoi_thuc_hien", "").strip()
            ghi_chu = request.form.get("ghi_chu", "").strip()

            # ===== VALIDATE =====
            if not so_dang_ky:
                return "Thiếu số đăng ký", 400

            if ngay_het_han < ngay_dang_ky:
                return "Ngày hết hạn không hợp lệ", 400

            # ===== LƯU LỊCH SỬ ĐĂNG KIỂM =====
            execute_retry(con, """
                INSERT INTO dang_kiem
                (vehicle_id, so_dang_ky, loai,
                 ngay_dang_ky, ngay_het_han,
                 trung_tam, chi_phi,
                 nguoi_thuc_hien, ghi_chu)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                vehicle_id,
                so_dang_ky,
                loai,
                ngay_dang_ky,
                ngay_het_han,
                trung_tam,
                chi_phi,
                nguoi_thuc_hien,
                ghi_chu
            ))

            # ===== CẬP NHẬT HẠN ĐĂNG KIỂM HIỆN TẠI CHO XE =====
            execute_retry(con, """
                UPDATE vehicles
                SET ngay_het_han_dang_kiem = ?
                WHERE id = ?
            """, (ngay_het_han, vehicle_id))

            con.commit()

        except Exception as e:
            con.rollback()
            return f"Lỗi hệ thống: {str(e)}", 500

        finally:
            con.close()

        return redirect("/lich-su-dang-kiem")

    # ===== LOAD DANH SÁCH XE =====
    vehicles = con.execute("""
        SELECT id, plate
        FROM vehicles
        ORDER BY plate
    """).fetchall()

    con.close()

    return render_template(
        "them_dang_kiem.html",
        vehicles=vehicles
    )
# =========================
# lịch sử bảo dưỡng
# =========================



@app.route("/lich-su-bao-duong")
@login_required
def lich_su_bao_duong():

    con = db()

    search = request.args.get("search", "")
    tu_ngay = request.args.get("tu_ngay")
    den_ngay = request.args.get("den_ngay")

    sql = """
        SELECT bd.*, v.plate
        FROM bao_duong bd
        JOIN vehicles v ON bd.vehicle_id = v.id
        WHERE 1=1
    """

    params = []

    if search:
        sql += " AND (v.plate LIKE ? OR bd.noi_dung LIKE ?)"
        params += [f"%{search}%", f"%{search}%"]

    if tu_ngay:
        sql += " AND bd.ngay_thuc_hien >= ?"
        params.append(tu_ngay)

    if den_ngay:
        sql += " AND bd.ngay_thuc_hien <= ?"
        params.append(den_ngay)

    sql += " ORDER BY bd.ngay_thuc_hien DESC"

    data = con.execute(sql, params).fetchall()

    # =========================
    # EXPORT EXCEL
    # =========================
    if request.args.get("export") == "excel":

        wb = Workbook()
        ws = wb.active
        ws.title = "BaoDuong"

        headers = [
            "Biển số",
            "Loại",
            "Nội dung",
            "Ngày thực hiện",
            "Ngày hoàn thành",
            "Trạng thái",
            "Đơn vị",
            "Chi phí",
            "KM tại thời điểm"
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for r in data:
            ws.append([
                r["plate"],
                r["loai"],
                r["noi_dung"],
                r["ngay_thuc_hien"],
                r["ngay_hoan_thanh"],
                r["trang_thai"],
                r["don_vi"],
                r["chi_phi"],
                r["km_tai_thoi_diem"]
            ])

        # Auto width
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return send_file(
            file_stream,
            download_name="lich_su_bao_duong.xlsx",
            as_attachment=True
        )

    con.close()

    return render_template(
        "lich_su_bao_duong.html",
        danh_sach=data
    )


# =========================
# thêm bảo dưỡng
# =========================
# =========================
# THÊM BẢO DƯỠNG (CHUẨN NGHIỆP VỤ)
# =========================
@app.route("/them-bao-duong", methods=["GET", "POST"])
def them_bao_duong():
    con = db()

    if request.method == "POST":
        try:
            vehicle_id = request.form["vehicle_id"]
            loai = request.form["loai"]
            noi_dung = request.form["noi_dung"].strip()
            ngay_thuc_hien = request.form["ngay_thuc_hien"]
            ngay_hoan_thanh = request.form.get("ngay_hoan_thanh")
            trang_thai = request.form["trang_thai"]
            don_vi = request.form.get("don_vi", "").strip()
            ghi_chu = request.form.get("ghi_chu", "").strip()

            # ===== XỬ LÝ CHI PHÍ =====
            chi_phi_raw = request.form.get("chi_phi", "0")
            chi_phi_clean = chi_phi_raw.replace(".", "").replace(",", "")
            try:
                chi_phi = int(chi_phi_clean)
                if chi_phi < 0:
                    return "Chi phí không hợp lệ", 400
            except:
                chi_phi = 0

            # ===== VALIDATE NGÀY =====
            if trang_thai == "hoan_thanh":
                if not ngay_hoan_thanh:
                    return "Thiếu ngày hoàn thành", 400
                if ngay_hoan_thanh < ngay_thuc_hien:
                    return "Ngày hoàn thành không hợp lệ", 400

            # ===== LẤY KM HIỆN TẠI CỦA XE =====
            vehicle = con.execute(
                "SELECT km FROM vehicles WHERE id=?",
                (vehicle_id,)
            ).fetchone()

            if not vehicle:
                return "Không tìm thấy xe", 404

            current_km = vehicle["km"] or 0

            # ===== 1️⃣ LƯU LỊCH SỬ BẢO DƯỠNG =====
            execute_retry(con, """
                INSERT INTO bao_duong
                (vehicle_id, loai, noi_dung,
                 ngay_thuc_hien, ngay_hoan_thanh,
                 trang_thai, don_vi, chi_phi,
                 km_tai_thoi_diem, ghi_chu)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                vehicle_id,
                loai,
                noi_dung,
                ngay_thuc_hien,
                ngay_hoan_thanh,
                trang_thai,
                don_vi,
                chi_phi,
                current_km,
                ghi_chu
            ))

            # ===== 2️⃣ CẬP NHẬT MỐC KM NẾU HOÀN THÀNH =====
            if trang_thai == "hoan_thanh":
                execute_retry(con, """
                    UPDATE vehicles
                    SET last_maintenance_km = ?
                    WHERE id = ?
                """, (current_km, vehicle_id))

            con.commit()

        except Exception as e:
            con.rollback()
            return f"Lỗi hệ thống: {str(e)}", 500

        finally:
            con.close()

        return redirect("/lich-su-bao-duong")

    # ===== LOAD DANH SÁCH XE =====
    vehicles = con.execute("""
        SELECT id, plate
        FROM vehicles
        ORDER BY plate
    """).fetchall()

    con.close()
    return render_template("them_bao_duong.html", vehicles=vehicles)


# =========================
# báo cáo km
# =========================

@app.route("/bao-cao-km")
def bao_cao_km():
    con = db()

    data = con.execute("""
        SELECT
            plate,
            strftime('%Y-%m', start_time) AS thang,
            SUM(km_travel) AS tong_km
        FROM trip_history
        GROUP BY plate, thang
        ORDER BY thang DESC
    """).fetchall()

    con.close()

    return render_template("bao_cao_km.html", data=data)
# =========================
# quản lý xe
# =========================

# =========================
# QUẢN LÝ XE
# =========================

from datetime import datetime, date

@app.route("/quan-ly-xe", methods=["GET", "POST"])
def quan_ly_xe():
    con = db()

    # =========================
    # THÊM XE
    # =========================
    if request.method == "POST":
        try:
            plate = request.form["plate"].strip()
            brand = request.form.get("brand")
            year = request.form.get("year")
            km = int(request.form.get("km") or 0)
            fuel_norm = float(request.form.get("fuel_norm") or 0)
            ngay_het_han_dk = request.form.get("ngay_het_han_dang_kiem")

            execute_retry(con, """
                INSERT INTO vehicles (
                    plate,
                    brand,
                    year,
                    km,
                    fuel_norm,
                    last_maintenance_km,
                    maintenance_cycle,
                    ngay_het_han_dang_kiem,
                    status,
                    driver_id,
                    start_time,
                    end_time,
                    work_content
                )
                VALUES (?, ?, ?, ?, ?, ?, 5000, ?, 
                        0, NULL, NULL, NULL, NULL)
            """, (
                plate,
                brand,
                year,
                km,
                fuel_norm,
                km,   # mốc bảo dưỡng = km hiện tại
                ngay_het_han_dk
            ))

            con.commit()

        except Exception as e:
            con.rollback()
            return f"Lỗi hệ thống: {str(e)}", 500

        finally:
            con.close()

        return redirect("/quan-ly-xe")

    # =========================
    # DASHBOARD ĐẾM SỐ LƯỢNG
    # =========================
    tong_xe = con.execute(
        "SELECT COUNT(*) FROM vehicles"
    ).fetchone()[0]

    qua_han_bd = con.execute("""
        SELECT COUNT(*) FROM vehicles
        WHERE (km - IFNULL(last_maintenance_km,0))
              >= IFNULL(maintenance_cycle,5000)
    """).fetchone()[0]

    qua_han_dk = con.execute("""
        SELECT COUNT(*) FROM vehicles
        WHERE ngay_het_han_dang_kiem IS NOT NULL
        AND DATE(ngay_het_han_dang_kiem) < DATE('now')
    """).fetchone()[0]

    # =========================
    # LỌC THEO DASHBOARD
    # =========================
    filter_type = request.args.get("filter")

    sql = """
        SELECT *
        FROM vehicles
        WHERE 1=1
    """

    if filter_type == "bao_duong":
        sql += """
        AND (km - IFNULL(last_maintenance_km,0))
            >= IFNULL(maintenance_cycle,5000)
        """

    elif filter_type == "dang_kiem":
        sql += """
        AND ngay_het_han_dang_kiem IS NOT NULL
        AND DATE(ngay_het_han_dang_kiem) < DATE('now')
        """

    sql += " ORDER BY plate"

    vehicles_raw = con.execute(sql).fetchall()

    # =========================
    # XỬ LÝ CẢNH BÁO CHI TIẾT
    # =========================
    today = date.today()
    vehicles = []

    for v in vehicles_raw:

        km_hien_tai = v["km"] or 0
        last_bd = v["last_maintenance_km"] or 0
        cycle = v["maintenance_cycle"] or 5000
        km_da_di = km_hien_tai - last_bd

        # Cảnh báo bảo dưỡng
        if km_da_di >= cycle:
            canh_bao_bd = "due"
        elif km_da_di >= cycle - 500:
            canh_bao_bd = "warning"
        else:
            canh_bao_bd = "ok"

        # Cảnh báo đăng kiểm
        canh_bao_dk = "ok"
        if v["ngay_het_han_dang_kiem"]:
            expiry = datetime.strptime(
                v["ngay_het_han_dang_kiem"], "%Y-%m-%d"
            ).date()

            diff = (expiry - today).days

            if diff < 0:
                canh_bao_dk = "expired"
            elif diff <= 30:
                canh_bao_dk = "warning"

        vehicles.append({
            **dict(v),
            "km_da_di": km_da_di,
            "canh_bao_bd": canh_bao_bd,
            "canh_bao_dk": canh_bao_dk
        })

    con.close()

    return render_template(
        "quan_ly_xe.html",
        vehicles=vehicles,
        tong_xe=tong_xe,
        qua_han_bd=qua_han_bd,
        qua_han_dk=qua_han_dk,
        filter_type=filter_type
    )
# =========================
# XÓA XE
# =========================
@app.route("/xoa-xe/<int:vid>")
def xoa_xe(vid):
    con = db()

    busy = con.execute("""
        SELECT 1 FROM vehicles
        WHERE id = ? AND status = 1
    """, (vid,)).fetchone()

    if busy:
        return "Không thể xóa: Xe đang được điều", 400

    execute_retry(con, "DELETE FROM vehicles WHERE id = ?", (vid,))
    con.commit()
    con.close()

    return redirect("/quan-ly-xe")


# =========================
# SỬA XE (HOÀN CHỈNH)
# =========================
@app.route("/sua-xe/<int:vid>", methods=["GET", "POST"])
def sua_xe(vid):
    con = db()

    if request.method == "POST":
        try:
            # ===== LẤY DỮ LIỆU TỪ FORM =====
            plate = request.form["plate"].strip()
            brand = request.form.get("brand", "").strip()
            year = request.form.get("year") or None
            km = int(request.form.get("km") or 0)
            fuel_norm = float(request.form.get("fuel_norm") or 0)

            last_maintenance_km = int(
                request.form.get("last_maintenance_km") or 0
            )

            maintenance_cycle = int(
                request.form.get("maintenance_cycle") or 5000
            )

            ngay_het_han_dk = request.form.get(
                "ngay_het_han_dang_kiem"
            ) or None

            # ===== VALIDATE NGHIỆP VỤ =====
            if km < 0:
                return "KM không hợp lệ", 400

            if last_maintenance_km > km:
                return "KM bảo dưỡng không thể lớn hơn KM hiện tại", 400

            if maintenance_cycle <= 0:
                maintenance_cycle = 5000

            # ===== UPDATE DATABASE =====
            execute_retry(con, """
                UPDATE vehicles
                SET plate=?,
                    brand=?,
                    year=?,
                    km=?,
                    fuel_norm=?,
                    last_maintenance_km=?,
                    maintenance_cycle=?,
                    ngay_het_han_dang_kiem=?
                WHERE id=?
            """, (
                plate,
                brand,
                year,
                km,
                fuel_norm,
                last_maintenance_km,
                maintenance_cycle,
                ngay_het_han_dk,
                vid
            ))

            con.commit()

        except Exception as e:
            con.rollback()
            return f"Lỗi hệ thống: {str(e)}", 500

        finally:
            con.close()

        return redirect("/quan-ly-xe")

    # =========================
    # GET: LẤY DỮ LIỆU XE
    # =========================
    vehicle = con.execute("""
        SELECT
            id,
            plate,
            brand,
            year,
            km,
            fuel_norm,
            last_maintenance_km,
            maintenance_cycle,
            ngay_het_han_dang_kiem
        FROM vehicles
        WHERE id=?
    """, (vid,)).fetchone()

    con.close()

    if not vehicle:
        return "Không tìm thấy xe", 404

    return render_template(
        "sua_xe.html",
        vehicle=vehicle
    )


# =========================
# LOGIN
# =========================

@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        username = request.form["username"].strip()
        password = request.form["password"]

        con = db()
        user = con.execute("""
            SELECT * FROM users
            WHERE username=? AND is_active=1
        """, (username,)).fetchone()
        con.close()

        if user and check_password_hash(user["password_hash"], password):

            session.clear()
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["role"] = user["role"]
            session["driver_id"] = user["driver_id"]
            role = user["role"]
            if user["role"] == "admin":
                return redirect("/")
            elif role == "driver":
                return redirect("/dieu-xe-driver")
            else:
                return redirect("/")   # user + viewer vào home

        return render_template("login_dieuxe.html", error="Sai tài khoản hoặc mật khẩu")

    return render_template("login_dieuxe.html")

# =========================
# LOOUT
# =========================



@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


# =========================
# dieu xe dirver
# =========================
@app.route("/dieu-xe-driver")
@login_required
@driver_required
def dieu_xe_driver():

    username = session.get("username")  # chính là số điện thoại
    con = db()

    # 1️⃣ Lấy driver theo số điện thoại
    driver = con.execute("""
        SELECT id, name
        FROM drivers
        WHERE phone = ?
    """, (username,)).fetchone()

    if not driver:
        con.close()
        return "Không tìm thấy tài xế tương ứng", 403

    # 2️⃣ Lấy xe đang công tác của tài xế đó
    vehicle = con.execute("""
        SELECT *
        FROM vehicles
        WHERE driver_id = ?
        AND status = 1
    """, (driver["id"],)).fetchone()

    con.close()

    return render_template(
        "dieu_xe_driver.html",
        vehicle=vehicle,
        driver_name=driver["name"]
    )
# =========================
# quan ly user
# =========================

@app.route("/quan-ly-user", methods=["GET", "POST"])
@login_required
@admin_required
def quan_ly_user():

    con = db()

    # =========================
    # THÊM USER
    # =========================
    if request.method == "POST":
        try:
            username = request.form["username"].strip()
            password = request.form["password"]
            role = request.form["role"]

            driver_id = request.form.get("driver_id") or None
            zalo_user_id = request.form.get("zalo_user_id") or None
            telegram_chat_id = request.form.get("telegram_chat_id") or None

            password_hash = generate_password_hash(password)

            execute_retry(con, """
                INSERT INTO users (
                    username, password_hash, role,
                    driver_id, zalo_user_id, telegram_chat_id, is_active
                )
                VALUES (?, ?, ?, ?, ?, ?, 1)
            """, (
                username,
                password_hash,
                role,
                driver_id,
                zalo_user_id,
                telegram_chat_id
            ))

            con.commit()

        except Exception as e:
            con.rollback()
            return f"Lỗi hệ thống: {str(e)}", 500

    # =========================
    # LẤY DANH SÁCH
    # =========================
    users = con.execute("""
        SELECT u.*, d.name
        FROM users u
        LEFT JOIN drivers d ON u.driver_id = d.id
        ORDER BY u.id DESC
    """).fetchall()

    drivers = con.execute("""
        SELECT id, name FROM drivers ORDER BY name
    """).fetchall()

    con.close()

    return render_template(
        "quan_ly_user.html",
        users=users,
        drivers=drivers
    )
# =========================
# sua user
# =========================
@app.route("/sua-user/<int:id>", methods=["GET", "POST"])
@login_required
@admin_required
def sua_user(id):

    con = db()

    if request.method == "POST":
        try:
            con.execute("""
                UPDATE users
                SET username=?,
                    role=?,
                    driver_id=?,
                    zalo_user_id=?,
                    telegram_chat_id=?
                WHERE id=?
            """, (
                request.form["username"],
                request.form["role"],
                request.form.get("driver_id") or None,
                request.form.get("zalo_user_id") or None,
                request.form.get("telegram_chat_id") or None,
                id
            ))

            con.commit()

        except Exception as e:
            con.rollback()
            return f"Lỗi hệ thống: {str(e)}", 500

        finally:
            con.close()

        return redirect("/quan-ly-user")

    user = con.execute("SELECT * FROM users WHERE id=?", (id,)).fetchone()
    drivers = con.execute("SELECT id, name FROM drivers").fetchall()

    con.close()

    return render_template("sua_user.html", row=user, drivers=drivers)
# =========================
# change password
# =========================

@app.route("/change_password", methods=["GET","POST"])
@login_required
def change_password():

    if request.method == "POST":
        old = request.form["old_password"]
        new = request.form["new_password"]

        con = db()
        user = con.execute("""
            SELECT * FROM users WHERE id=?
        """, (session["user_id"],)).fetchone()

        if not check_password_hash(user["password_hash"], old):
            return render_template("change_password.html",
                                   error="Mật khẩu cũ không đúng")

        con.execute("""
            UPDATE users
            SET password_hash=?
            WHERE id=?
        """, (generate_password_hash(new), session["user_id"]))

        con.commit()
        con.close()

        return redirect("/")

    return render_template("change_password.html")

# =========================
# Sửa bảo dưỡng
# =========================


@app.route("/sua-bao-duong/<int:id>", methods=["GET","POST"])
@login_required
@admin_required
def sua_bao_duong(id):

    con = db()

    if request.method == "POST":
        con.execute("""
            UPDATE bao_duong
            SET loai=?, noi_dung=?, ngay_thuc_hien=?,
                ngay_hoan_thanh=?, trang_thai=?,
                don_vi=?, chi_phi=?, ghi_chu=?
            WHERE id=?
        """, (
            request.form["loai"],
            request.form["noi_dung"],
            request.form["ngay_thuc_hien"],
            request.form.get("ngay_hoan_thanh"),
            request.form["trang_thai"],
            request.form.get("don_vi"),
            request.form.get("chi_phi") or 0,
            request.form.get("ghi_chu"),
            id
        ))

        con.commit()
        con.close()
        return redirect("/lich-su-bao-duong")

    data = con.execute(
        "SELECT * FROM bao_duong WHERE id=?",
        (id,)
    ).fetchone()

    con.close()

    return render_template("sua_bao_duong.html", data=data)

# =========================
# Xóa bảo dưỡng
# =========================


@app.route("/xoa-bao-duong/<int:id>")
@login_required
@admin_required
def xoa_bao_duong(id):

    con = db()
    execute_retry(con, "DELETE FROM bao_duong WHERE id=?", (id,))
    con.commit()
    con.close()

    return redirect("/lich-su-bao-duong")

# =========================
# Sửa đăng kiểm
# =========================


@app.route("/sua-dang-kiem/<int:id>", methods=["GET","POST"])
@login_required
@admin_required
def sua_dang_kiem(id):

    con = db()

    if request.method == "POST":
        con.execute("""
            UPDATE dang_kiem
            SET so_dang_ky=?, loai=?, ngay_dang_ky=?,
                ngay_het_han=?, trung_tam=?, chi_phi=?, ghi_chu=?
            WHERE id=?
        """, (
            request.form["so_dang_ky"],
            request.form["loai"],
            request.form["ngay_dang_ky"],
            request.form["ngay_het_han"],
            request.form.get("trung_tam"),
            request.form.get("chi_phi") or 0,
            request.form.get("ghi_chu"),
            id
        ))

        con.commit()
        con.close()
        return redirect("/lich-su-dang-kiem")

    data = con.execute(
        "SELECT * FROM dang_kiem WHERE id=?",
        (id,)
    ).fetchone()

    con.close()

    return render_template("sua_dang_kiem.html", data=data)

# =========================
# Xóa đăng kiểm
# =========================

@app.route("/xoa-dang-kiem/<int:id>")
@login_required
@admin_required
def xoa_dang_kiem(id):

    con = db()
    execute_retry(con, "DELETE FROM dang_kiem WHERE id=?", (id,))
    con.commit()
    con.close()

    return redirect("/lich-su-dang-kiem")



# =========================
# THỐNG KÊ
# =========================
@app.route("/thong-ke")
@login_required
@admin_required
def thong_ke():

    con = db()

    danh_sach_xe = con.execute("""
        SELECT DISTINCT plate FROM trip_history ORDER BY plate
    """).fetchall()

    danh_sach_taixe = con.execute("""
        SELECT DISTINCT driver_name FROM trip_history ORDER BY driver_name
    """).fetchall()

    tu_ngay = request.args.get("tu_ngay")
    den_ngay = request.args.get("den_ngay")
    xe = request.args.get("xe")
    taixe = request.args.get("taixe")
    export = request.args.get("export")

    ket_qua = []
    co_dieu_kien = any([tu_ngay, den_ngay, xe, taixe])

    tong_km = 0
    tong_chuyen = 0
    tong_phut = 0

    bieu_do = defaultdict(lambda: {"km": 0, "chuyen": 0})

    if co_dieu_kien:

        sql = """
            SELECT plate, driver_name, work_content,
                   start_time, end_time,
                   duration_minutes,
                   km_travel
            FROM trip_history
            WHERE 1=1
        """
        params = []

        if tu_ngay:
            sql += " AND DATE(start_time) >= ?"
            params.append(tu_ngay)

        if den_ngay:
            sql += " AND DATE(start_time) <= ?"
            params.append(den_ngay)

        if xe:
            sql += " AND plate = ?"
            params.append(xe)

        if taixe:
            sql += " AND driver_name = ?"
            params.append(taixe)

        sql += " ORDER BY start_time DESC"

        ket_qua = con.execute(sql, params).fetchall()

        # ================================
        # TÍNH TỔNG HỢP
        # ================================
        for r in ket_qua:

            km = r["km_travel"] or 0
            phut = r["duration_minutes"] or 0

            tong_km += km
            tong_phut += phut
            tong_chuyen += 1

            thang = r["start_time"][:7]  # YYYY-MM
            bieu_do[thang]["km"] += km
            bieu_do[thang]["chuyen"] += 1

        # ================================
        # XUẤT EXCEL
        # ================================
        if export == "excel" and ket_qua:

            wb = Workbook()
            ws = wb.active
            ws.title = "ThongKeDieuXe"

            headers = [
                "Xe",
                "Tài xế",
                "Nội dung công tác",
                "Giờ đi",
                "Giờ về",
                "KM",
                "Thời gian (phút)"
            ]

            ws.append(headers)

            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)

            for r in ket_qua:
                ws.append([
                    r["plate"],
                    r["driver_name"],
                    r["work_content"],
                    format_datetime(r["start_time"]),
                    format_datetime(r["end_time"]),
                    r["km_travel"] or 0,
                    r["duration_minutes"] or 0
                ])

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            con.close()

            return send_file(
                output,
                download_name="thong_ke_dieu_xe.xlsx",
                as_attachment=True,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # ================================
    # SAU KHI XỬ LÝ XONG
    # ================================
    con.close()

    tong_gio = round(tong_phut / 60, 2)

    thang_labels = sorted(bieu_do.keys())
    data_km = [bieu_do[t]["km"] for t in thang_labels]
    data_chuyen = [bieu_do[t]["chuyen"] for t in thang_labels]

    return render_template(
        "thong_ke.html",
        danh_sach_xe=danh_sach_xe,
        danh_sach_taixe=danh_sach_taixe,
        ket_qua=ket_qua,
        tu_ngay=tu_ngay,
        den_ngay=den_ngay,
        xe=xe,
        taixe=taixe,
        co_dieu_kien=co_dieu_kien,
        tong_km=tong_km,
        tong_chuyen=tong_chuyen,
        tong_gio=tong_gio,
        thang_labels=thang_labels,
        data_km=data_km,
        data_chuyen=data_chuyen
    )

    # ================================
    # lấy webhook
    # ================================


# =========================
# ZALO BOT WEBHOOK
# =========================
@app.route("/zalo-webhook", methods=["POST"])
def zalo_webhook():

    secret_header = request.headers.get("X-Bot-Api-Secret-Token")
    my_secret = os.getenv("ZALO_SECRET_TOKEN")

    if secret_header != my_secret:
        return "Unauthorized", 403

    data = request.json

    if data.get("event_name") != "message.text.received":
        return "OK"

    user_id = data["message"]["from"]["id"]
    text = data["message"]["text"].strip().lower()

    print("User:", user_id)
    print("Raw text:", text)

    con = db()

    # =========================
    # LỆNH KETNOI
    # =========================
    if text.startswith("ketnoi"):

        parts = text.split(" ")

        if len(parts) < 2:

            gui_zalo_cho_taixe(
                user_id,
                "📱 Cú pháp đúng:\nketnoi 0905086253"
            )

            con.close()
            return "OK"

        phone = parts[1]

        # chuẩn hóa số điện thoại
        phone = phone.replace(" ", "").replace(".", "").replace("-", "")

        print("Phone cleaned:", phone)

        driver = con.execute("""
            SELECT id,name
            FROM drivers
            WHERE phone = ?
        """, (phone,)).fetchone()

        if driver:

            con.execute("""
                UPDATE drivers
                SET zalo_user_id = ?
                WHERE id = ?
            """, (user_id, driver["id"]))

            con.commit()

            gui_zalo_cho_taixe(
                user_id,
                "✅ Đã liên kết Zalo với hệ thống điều xe."
            )

            print("Đã cập nhật Zalo ID cho:", driver["name"])

        else:

            gui_zalo_cho_taixe(
                user_id,
                "❌ Số điện thoại chưa có trong hệ thống."
            )

        con.close()
        return "OK"

    # =========================
    # TRƯỜNG HỢP GỬI SỐ ĐIỆN THOẠI TRỰC TIẾP
    # =========================

    phone = text.replace(" ", "").replace(".", "").replace("-", "")

    if phone.isdigit():

        driver = con.execute("""
            SELECT id,name
            FROM drivers
            WHERE phone = ?
        """, (phone,)).fetchone()

        if driver:

            con.execute("""
                UPDATE drivers
                SET zalo_user_id = ?
                WHERE id = ?
            """, (user_id, driver["id"]))

            con.commit()

            gui_zalo_cho_taixe(
                user_id,
                "✅ Đã liên kết Zalo với hệ thống điều xe."
            )

            print("Đã cập nhật Zalo ID cho:", driver["name"])

        else:

            gui_zalo_cho_taixe(
                user_id,
                "❌ Số điện thoại chưa có trong hệ thống."
            )

    else:

        gui_zalo_cho_taixe(
            user_id,
            "📱 Gửi số điện thoại hoặc:\nketnoi 0905086253"
        )

    con.close()

    return "OK"


# =========================
# TELEGRAM BOT WEBHOOK
# =========================
@app.route("/telegram-webhook", methods=["POST"])
def telegram_webhook():

    try:

        data = request.get_json(force=True)

        print("Telegram update:", data)

        if "message" not in data:
            return "OK"

        message = data["message"]

        if "text" not in message:
            return "OK"

        chat_id = message["chat"]["id"]
        text = message["text"].strip().lower()
        ai = ai_parse_command(text)

        print("Telegram chat_id:", chat_id)
        print("Raw text:", text)

        con = db()

        # =========================
        # KIỂM TRA ADMIN
        # =========================
        admin = con.execute("""
            SELECT id FROM bot_admins WHERE telegram_id=?
        """, (chat_id,)).fetchone()

        is_admin = True if admin else False

        # =========================
        # /dsxe
        # =========================
        if text == "/dsxe":

            if not is_admin:
                send_telegram(chat_id, "❌ Bạn không có quyền.")
                con.close()
                return "OK"

            rows = con.execute("""
                SELECT v.plate, d.name, v.work_content
                FROM vehicles v
                LEFT JOIN drivers d ON v.driver_id = d.id
                WHERE v.status = 1
            """).fetchall()

            if not rows:
                send_telegram(chat_id, "🚗 Hiện không có xe nào đang hoạt động.")
                con.close()
                return "OK"

            msg = "🚗 DANH SÁCH XE ĐANG HOẠT ĐỘNG\n\n"

            for r in rows:
                msg += (
                    f"Xe: {r['plate']}\n"
                    f"Tài xế: {r['name']}\n"
                    f"Nội dung: {r['work_content']}\n\n"
                )

            send_telegram(chat_id, msg)
            con.close()
            return "OK"

        # =========================
        # /dsxeranh
        # =========================
        if text == "/dsxeranh":

            if not is_admin:
                send_telegram(chat_id, "❌ Bạn không có quyền.")
                con.close()
                return "OK"

            rows = con.execute("""
                SELECT plate FROM vehicles WHERE status = 0
            """).fetchall()

            if not rows:
                send_telegram(chat_id, "🚗 Hiện không có xe nào rãnh.")
                con.close()
                return "OK"

            msg = "🚗 DANH SÁCH XE RÃNH\n\n"

            for r in rows:
                msg += f"Xe: {r['plate']}\n"

            send_telegram(chat_id, msg)
            con.close()
            return "OK"

        # =========================
        # /taixeranh
        # =========================
        if text == "/taixeranh":

            if not is_admin:
                send_telegram(chat_id, "❌ Bạn không có quyền.")
                con.close()
                return "OK"

            rows = con.execute("""
                SELECT d.name, d.phone
                FROM drivers d
                WHERE d.id NOT IN (
                    SELECT driver_id
                    FROM vehicles
                    WHERE status = 1 AND driver_id IS NOT NULL
                )
                ORDER BY d.name
            """).fetchall()

            if not rows:
                send_telegram(chat_id, "⚠️ Không có tài xế rảnh.")
                con.close()
                return "OK"

            msg = "👨‍✈️ DANH SÁCH TÀI XẾ RẢNH\n\n"

            for r in rows:
                msg += f"Tài xế: {r['name']}\nSĐT: {r['phone']}\n\n"

            send_telegram(chat_id, msg)
            con.close()
            return "OK"

        # =========================
        # /taixe
        # =========================
        if text == "/taixe":

            rows = con.execute("""
                SELECT d.name, v.plate
                FROM drivers d
                LEFT JOIN vehicles v
                ON v.driver_id=d.id AND v.status=1
            """).fetchall()

            msg = "👨‍✈️ TÀI XẾ\n\n"

            for r in rows:
                if r["plate"]:
                    msg += f"{r['name']} → 🚗 {r['plate']}\n"
                else:
                    msg += f"{r['name']} → rảnh\n"

            send_telegram(chat_id, msg)
            con.close()
            return "OK"

        # =========================
        # /thongke
        # =========================
        if text == "/thongke":

            xe_chay = con.execute("SELECT COUNT(*) FROM vehicles WHERE status=1").fetchone()[0]
            xe_ranh = con.execute("SELECT COUNT(*) FROM vehicles WHERE status=0").fetchone()[0]

            taixe_ranh = con.execute("""
                SELECT COUNT(*) FROM drivers
                WHERE id NOT IN (
                    SELECT driver_id FROM vehicles
                    WHERE status=1 AND driver_id IS NOT NULL
                )
            """).fetchone()[0]

            msg = f"""
📊 THỐNG KÊ HỆ THỐNG

🚗 Xe đang chạy: {xe_chay}
🚗 Xe rảnh: {xe_ranh}
👨‍✈️ Tài xế rảnh: {taixe_ranh}
"""

            send_telegram(chat_id, msg)
            con.close()
            return "OK"

        # =========================
        # KHÔNG NHẬN DIỆN
        # =========================
        send_telegram(
            chat_id,
            "📱 Lệnh hỗ trợ:\n"
            "/dsxe\n/dsxeranh\n/taixeranh\n"
            "/xe 94A-001.88\n/taixe\n/thongke\nketnoi 0905xxxx"
        )

        con.close()
        return "OK"

    except Exception as e:
        print("Telegram webhook error:", e)
        return "OK"


# =========================
# Dashboard realtime xe đang chạy
# =========================
@app.route("/dashboard-data")
def dashboard_data():

    con = db()

    rows = con.execute("""
        SELECT v.plate, d.name, v.work_content, v.start_time
        FROM vehicles v
        LEFT JOIN drivers d ON v.driver_id = d.id
        WHERE v.status = 1
    """).fetchall()

    data = []

    for r in rows:
        data.append({
            "xe": r["plate"],
            "taixe": r["name"],
            "noidung": r["work_content"],
            "start": r["start_time"]
        })

    con.close()

    return {"data": data}
# =========================
# yêu cầu điều xe
# =========================
@app.route("/yeu-cau-dieu-xe", methods=["GET", "POST"])
@login_required
def yeu_cau_dieu_xe():

    con = db()

    # =========================
    # THÊM YÊU CẦU
    # =========================
    if request.method == "POST":
        execute_retry(con, """
            INSERT INTO yeu_cau_xe (
                nguoi_yeu_cau, chuc_vu, so_hanh_khach,
                muc_dich, diem_don, diem_den,
                ngay_di, ngay_ve,
                trang_thai, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'cho_duyet', datetime('now'))
        """, (
            request.form["nguoi_yeu_cau"],
            request.form.get("chuc_vu"),
            request.form.get("so_hanh_khach"),
            request.form.get("muc_dich"),
            request.form.get("diem_don"),
            request.form.get("diem_den"),
            request.form.get("ngay_di"),
            request.form.get("ngay_ve")
        ))
        con.commit()

    # =========================
    # FILTER
    # =========================
    status = request.args.get("status")

    sql = "SELECT * FROM yeu_cau_xe WHERE 1=1"

    if status == "cho_duyet":
        sql += " AND trang_thai='cho_duyet'"
    elif status == "da_duyet":
        sql += " AND trang_thai='da_duyet'"

    sql += " ORDER BY created_at DESC"

    data_raw = con.execute(sql).fetchall()

    data = []
    for r in data_raw:
        r = dict(r)

        r["ngay_di_dep"] = format_datetime(r.get("ngay_di"))
        r["ngay_ve_dep"] = format_datetime(r.get("ngay_ve"))

        data.append(r)   # ✅ FIX LỖI 500

    # =========================
    # DASHBOARD
    # =========================
    tong = con.execute("SELECT COUNT(*) FROM yeu_cau_xe").fetchone()[0]

    cho = con.execute("""
        SELECT COUNT(*) FROM yeu_cau_xe 
        WHERE trang_thai='cho_duyet'
    """).fetchone()[0]

    da = con.execute("""
        SELECT COUNT(*) FROM yeu_cau_xe 
        WHERE trang_thai='da_duyet'
    """).fetchone()[0]

    con.close()

    return render_template(
        "yeu_cau_dieu_xe.html",
        data=data,
        tong=tong,
        cho=cho,
        da=da
    )
# =========================
# danh sách yêu cầu
# =========================
@app.route("/danh-sach-yeu-cau")
@login_required
def danh_sach_yeu_cau():

    con = db()

    status = request.args.get("status")
    search_name = request.args.get("search_name", "").strip()
    tu_ngay = request.args.get("tu_ngay")
    den_ngay = request.args.get("den_ngay")
    sql = "SELECT * FROM yeu_cau_xe WHERE 1=1"
    params = []
    if status == "cho_duyet":
        sql += " AND trang_thai='cho_duyet'"
    elif status == "da_duyet":
        sql += " AND trang_thai='da_duyet'"
    # lọc theo tên
    if search_name:
         sql += " AND nguoi_yeu_cau LIKE ?"
         params.append(f"%{search_name}%")

    # lọc từ ngày
    if tu_ngay:
        sql += " AND date(substr(ngay_di,1,10)) >= date(?)"
        params.append(tu_ngay)

    # lọc đến ngày
    if den_ngay:
        sql += " AND date(substr(ngay_di,1,10)) <= date(?)"
        params.append(den_ngay)
    sql += " ORDER BY id DESC"

    data_raw = con.execute(sql, params).fetchall()

    data = []
    for r in data_raw:
        r = dict(r)

        r["ngay_di_dep"] = format_date(r.get("ngay_di"))
        r["ngay_ve_dep"] = format_date(r.get("ngay_ve"))

        data.append(r)

    # 👉 lấy xe rảnh
    vehicles = con.execute("""
        SELECT id, plate
        FROM vehicles
        WHERE status = 0
    """).fetchall()

    # 👉 tài xế rảnh
    drivers = con.execute("""
        SELECT id, name
        FROM drivers
        WHERE id NOT IN (
            SELECT driver_id FROM vehicles
            WHERE status=1 AND driver_id IS NOT NULL
        )
    """).fetchall()
    tong = con.execute("SELECT COUNT(*) FROM yeu_cau_xe").fetchone()[0]
    cho = con.execute("""
    SELECT COUNT(*) FROM yeu_cau_xe 
    WHERE trang_thai='cho_duyet'
    """).fetchone()[0]

    da = con.execute("""
    SELECT COUNT(*) FROM yeu_cau_xe 
    WHERE trang_thai='da_duyet'
    """).fetchone()[0]
    con.close()

    return render_template(
        "danh_sach_yeu_cau.html",
        data=data,
        vehicles=vehicles,
        drivers=drivers,
        tong=tong,
        cho=cho,
        da=da,
        search_name=search_name,
        tu_ngay=tu_ngay,
        den_ngay=den_ngay
    )

# =========================
# xử lý yêu cầu
# =========================

@app.route("/xu-ly-yeu-cau/<int:id>", methods=["POST"])
@login_required
@admin_required
def xu_ly_yeu_cau(id):

    con = db()

    yc = con.execute("""
        SELECT * FROM yeu_cau_xe WHERE id=?
    """, (id,)).fetchone()
    yc = dict(yc)
    if not yc:
        return "Không tìm thấy yêu cầu", 404

    vehicle_id = request.form.get("vehicle_id")
    driver_id = request.form.get("driver_id")

    if not vehicle_id or not driver_id:
        return "Thiếu xe hoặc tài xế", 400

    # =========================
    # CẬP NHẬT XE
    # =========================
    start_time = yc["ngay_di"] or datetime.now().isoformat()
    end_time = yc["ngay_ve"] or datetime.now().isoformat()
    requester = yc["nguoi_yeu_cau"]
    work_content = f"{yc['muc_dich']}"

    execute_retry(con, """
        UPDATE vehicles
        SET status=1,
            driver_id=?,
            start_time=?,
            end_time=?,
            work_content=?,
            requester=?
        WHERE id=?
    """, (driver_id, start_time, end_time, work_content, requester, vehicle_id))

    # =========================
    # CẬP NHẬT YÊU CẦU
    # =========================
    execute_retry(con, """
        UPDATE yeu_cau_xe
        SET trang_thai='da_duyet'
        WHERE id=?
    """, (id,))

    
    info = con.execute("""
         SELECT v.plate,
                d.name,
                d.zalo_user_id,
                d.telegram_chat_id
         FROM vehicles v
         JOIN drivers d ON v.driver_id = d.id
         WHERE v.id = ?
    """, (vehicle_id)).fetchone()
    con.commit()
    con.close()
    # =========================
    # 🔥 GỬI ZALO + TELEGRAM
    # =========================
    try:
         ngaydi_dt = datetime.fromisoformat(yc["ngay_di"])
         ngay_di_dep = ngaydi_dt.strftime("%H:%M ngày %d/%m/%Y")
    except:
         ngay_di_dep = yc["ngay_di"]

    try:
         ngayve_dt = datetime.fromisoformat(yc["ngay_ve"])
         ngay_ve_dep = ngayve_dt.strftime("%H:%M ngày %d/%m/%Y")
    except:
         ngay_ve_dep = yc["ngay_ve"]
    if info:

        noi_dung = f"""
🚗 ĐIỀU XE 

Xe: {info['plate']}
Tài xế: {info['name']}
Thời gian đi: {ngay_di_dep}
Thời gian về: {ngay_ve_dep}
Người đi công tác: {yc['nguoi_yeu_cau']} 
Nội dung:
{yc['muc_dich']}
"""

        if info["zalo_user_id"]:
            gui_zalo_cho_taixe(info["zalo_user_id"], noi_dung)

        if info["telegram_chat_id"]:
            send_telegram(info["telegram_chat_id"], noi_dung)


    return redirect("/danh-sach-yeu-cau")


# =========================
# SỬA YÊU CẦU ĐIỀU XE
# =========================
@app.route("/sua-yeu-cau/<int:id>", methods=["GET", "POST"])
@login_required
def sua_yeu_cau(id):

    con = db()

    if request.method == "POST":
        con.execute("""
            UPDATE yeu_cau_xe
            SET nguoi_yeu_cau=?,
                chuc_vu=?,
                so_hanh_khach=?,
                ngay_di=?,
                ngay_ve=?,
                muc_dich=?,
                diem_don=?,
                diem_den=?
            WHERE id=?
        """, (
            request.form["nguoi_yeu_cau"],
            request.form.get("chuc_vu"),
            request.form.get("so_hanh_khach"),
            request.form.get("ngay_di"),
            request.form.get("ngay_ve"),
            request.form.get("muc_dich"),
            request.form.get("diem_don"),
            request.form.get("diem_den"),
            id
        ))
        con.execute("""
        UPDATE vehicles
        SET requester = ?, 
            work_content = ?
        WHERE yeu_cau_id = ?
        """)
        con.commit()
        con.close()

        return redirect("/danh-sach-yeu-cau")

    data = con.execute(
        "SELECT * FROM yeu_cau_xe WHERE id=?",
        (id,)
    ).fetchone()

    con.close()
    
    return render_template("sua_yeu_cau.html", data=data)
# =========================
# XÓA YÊU CẦU
# =========================
@app.route("/xoa-yeu-cau/<int:id>")
@login_required
def xoa_yeu_cau(id):

    con = db()

    # không cho xóa nếu đã điều xe
    check = con.execute("""
        SELECT trang_thai FROM yeu_cau_xe WHERE id=?
    """, (id,)).fetchone()

    if check and check["trang_thai"] != "cho_duyet":
        return "Không thể xóa: yêu cầu đã duyệt", 400

    execute_retry(con, "DELETE FROM yeu_cau_xe WHERE id=?", (id,))
    con.commit()
    con.close()

    return redirect("/danh-sach-yeu-cau")


# ================= XÓA USER =================

@app.route("/xoa-user/<int:id>")
@login_required
@admin_required
def xoa_user(id):

    con = db()
    execute_retry(con, "DELETE FROM users WHERE id=?", (id,))
    con.commit()
    con.close()

    return redirect("/quan-ly-user")

# =========================
# sao lưu dữ liệu
# =========================
@app.route("/backup")
def backup():

    return send_file(
        "fleet.db",
        as_attachment=True,
        download_name="fleet_backup.db"
    )
# =========================
# chống ngủ server
# =========================

@app.route("/ping")
def ping():
    return {"status": "ok"}, 200

# =========================
# chống ngủ server
# =========================
@app.route("/health")
def health():
    return "healthy", 200
# =========================
# chống ngủ server
# =========================
def keep_alive():
    while True:
        try:
            requests.get("https://quanly-xe.onrender.com/ping", timeout=5)
        except:
            pass
        time.sleep(300)

threading.Thread(target=keep_alive, daemon=True).start()
# =========================
# CHẠY APP
# =========================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))

    app.run(host="0.0.0.0", port=port, debug=False)
